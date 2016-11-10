import logging
import os
from collections import defaultdict

from openpyxl import load_workbook

from gaconverter.constants import ALLELE_COLUMNS_START, NUMBERS_COLUMN


class Extractor(object):
    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path

    def run(self):
        if not self._validate():
            return
        records_dicts = self._process_workbook()
        for allele, vals in records_dicts.items():
            print allele
            for sample, num in vals:
                print "--"
                print sample
                print num
            print "++++++++++++"

    def _process_workbook(self):
        records_dict = defaultdict(list)
        allele_column = {}
        number_row = {}
        wb = load_workbook(self.xlsx_path)

        ws = wb.active
        prev_name = ""
        for i in range(ALLELE_COLUMNS_START, ws.max_column + 1):
            val = ws.cell(row=1, column=i).value
            if val:
                if prev_name == val.lower():
                    continue
                allele_column[val.lower()] = i
                prev_name = val.lower()
            else:
                prev_name = ""
        for i in range(1, ws.max_row + 1):
            val = ws.cell(row=i, column=NUMBERS_COLUMN).value
            number_row[val] = i

        for allele_name, allele_n in allele_column.items():
            for sample_name, row_n in number_row.items():
                # print sample_name, row_n
                cell_left = ws.cell(row=row_n, column=allele_n)
                if not cell_left.value:
                    continue

                cell_right = ws.cell(row=row_n, column=allele_n + 1)
                if not cell_right.value:
                    records_dict[allele_name].append((sample_name, cell_left.value))
        return records_dict

    def _validate(self):
        if not os.path.exists(self.xlsx_path):
            logging.error("Cannot find xlsx file: %s!" % self.rtf_path)
            return False
        return True
