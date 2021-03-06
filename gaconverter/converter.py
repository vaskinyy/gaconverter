import logging
import os
from collections import defaultdict

from openpyxl import load_workbook
from pyth.plugins.plaintext.writer import PlaintextWriter

from gaconverter.constants import ALLELE_COLUMNS_START, NUMBERS_COLUMN


class GARecord(object):
    def __init__(self):
        self.number = None
        self.allele = None
        self.left = None
        self.right = None

    def valid(self):
        return self.number and self.allele and self.left

    def __str__(self):
        res = "{} {} {}".format(self.number, self.allele, self.left)
        if self.right:
            res += " {}".format(self.right)
        return res


class Converter(object):
    def __init__(self, txt_path, xlsx_path):
        self.txt_path = txt_path
        self.xlsx_path = xlsx_path

    def run(self):
        if not self._validate():
            return
        records = self._parse_data()
        records_dict = defaultdict(list)
        for record in records:
            records_dict[record.number].append(record)

        self._process_workbook(records_dict)
        logging.info("Processed {} items".format(len(records_dict)))

    def _parse_data(self):
        logging.info("Parsing GA file {}".format(self.txt_path))
        with open(self.txt_path, "r") as txt_file:
            for line in txt_file:
                line = line.strip()
                if not line:
                    continue
                columns = line.split('\t')
                if len(columns) < 3:
                    continue
                record = GARecord()
                try:
                    record.number = int(columns[1].split(' ')[2])
                except IndexError:
                    continue
                record.allele = columns[2]
                record.left = columns[3]
                if len(columns) >= 4:
                    record.right = columns[4]
                if not record.valid():
                    logging.warning("Skipping invalid record {}".format(line))
                    continue
                yield record

    def _process_workbook(self, records_dict):
        allele_column = {}
        number_row = {}
        wb = load_workbook(self.xlsx_path)

        ws = wb.active
        prev_name = ""
        for i in range(ALLELE_COLUMNS_START, ws.max_column + 1):
            val = ws.cell(row=1, column=i).value
            if val:
                if prev_name == val.lower():
                    continue
                allele_column[val.lower()] = i
                prev_name = val.lower()
            else:
                prev_name = ""
        for i in range(1, ws.max_row + 1):
            val = ws.cell(row=i, column=NUMBERS_COLUMN).value
            number_row[val] = i
        for number, records in records_dict.items():
            if number not in number_row:
                logging.warning(
                    "Can't find experiment number {} in xlsx file, please, add it".format(number))
                continue
            row = number_row[number]

            for record in records:
                if record.allele.lower() not in allele_column:
                    logging.warning(
                        "Can't find allele {} of experiment number {} in xlsx file, please, add it".format(
                            record.allele, number))
                    continue
                column = allele_column[record.allele.lower()]

                cell_left = ws.cell(row=row, column=column)
                if cell_left.value:
                    logging.warning(
                        "Data already inserted for {} of experiment number {}. Skipping".format(
                            record.allele, number))
                    continue
                cell_left.value = record.left

                if record.right:
                    cell_right = ws.cell(row=row, column=column + 1)
                    if cell_right.value:
                        logging.warning(
                            "Data already inserted for {} of experiment number {}. Skipping".format(
                                record.allele, number))
                        continue

                    cell_right.value = record.right

        wb.save(self.xlsx_path[:-5] + "_new" + ".xlsx")

    def _validate(self):
        if not os.path.exists(self.txt_path):
            logging.error("Cannot find txt file: %s!" % self.txt_path)
            return False

        if not os.path.exists(self.xlsx_path):
            logging.error("Cannot find xlsx file: %s!" % self.txt_path)
            return False
        return True
